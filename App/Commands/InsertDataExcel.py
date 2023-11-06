from sqlalchemy import create_engine, Table, Column, Text, Integer, Date, MetaData, func, text, update
from sqlalchemy_utils import database_exists, create_database
from sqlalchemy.orm import sessionmaker, declarative_base

import openpyxl
from openpyxl.styles import PatternFill, Color

from datetime import datetime

import asyncio
import re
import os


class InsertExcel:

    def __init__(self, mainPath) -> None: 
        self.excelFilePath =  os.path.abspath(os.path.join(mainPath, "..", "Data", "Excel", "Products.xlsx"))
        self.productDatabasePath = os.path.abspath(os.path.join(mainPath, "..", "Data", "Database", "ProductDatabase.db"))
        
        self.productEngine = create_engine("sqlite:///" + self.productDatabasePath)

        self.MainHandler()

    def MainHandler(self):
        self.createOrOpenExcelFile()
        self.getProductDatabaseLength()
        
        self.startedTime = datetime.now()
        self.writeToExcel()

    def createWorksheetTemplate(self):
        worksheet = self.wb.active
        worksheet.title = "시트1"

        header_row = ['brand', 'name', 'brand+name', 'brand+name without ()', 'link', 'inside ()']
        worksheet.append(header_row)
        
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for col_num, header in enumerate(header_row, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.fill = fill

        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 35
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 20
        worksheet.column_dimensions['G'].width = 15

    def createOrOpenExcelFile(self):
        if os.path.isfile(self.excelFilePath):
            self.wb = openpyxl.load_workbook(self.excelFilePath)
        else:
            self.wb = openpyxl.Workbook()
            self.createWorksheetTemplate()
            self.wb.save(self.excelFilePath)


    def getProductsDatas(self):
        Session = sessionmaker(bind=self.productEngine)
        Base = declarative_base()

        Base.metadata.create_all(self.productEngine)

        session = Session()
        query = text("SELECT * FROM ProductInf WHERE isProcessed = 0 LIMIT 100")
        results = session.execute(query).fetchall()
        session.close()

        return [result for result in results]

    def setProductToProcessed(self, dbID):
        Session = sessionmaker(bind=self.productEngine)
        Base = declarative_base()

        Base.metadata.create_all(self.productEngine)

        class ProductInf(declarative_base()):
            __tablename__ = 'ProductInf'

            id = Column(Integer, primary_key=True)
            brand = Column(Text)
            name = Column(Text)
            url = Column(Text)
            brackets = Column(Text)
            isProcessed = Column(Integer)
            processDate = Column(Date)
        
        session = Session()
        update_product = update(ProductInf).where(ProductInf.id == dbID).values(isProcessed=1, processDate=datetime.now())
        
        session.execute(update_product)
        session.commit()
        session.close()

    def organizeProductData(self, productDatas):
        result = []
        for product in productDatas:            
            productBrand = product.brand
            productName = product.name
            productBrandAndName = f"{productBrand} {productName}"
            productLink = product.url
            productBrandAndNameWithoutBrackets = re.sub(r'[()]', '', productBrandAndName)
            bracketContents = re.findall("\(([^)]*)\)", product.brackets)
            
            result.append([[productBrand, productName, productBrandAndName, productBrandAndNameWithoutBrackets, productLink], bracketContents])
            self.setProductToProcessed(product.id)
        return result
    
    def writeToExcel(self):
        while True:
            productsData = self.getProductsDatas()
            
            if not productsData:
                break

            def flatten(l):
                for el in l:
                    if isinstance(el, list):
                        yield from flatten(el)
                    else:
                        yield el

            organizedData = self.organizeProductData(productsData)
            
            ws = self.wb.active
            for data in organizedData:
                flatList = flatten(data)
                ws.append(flatList)
                
                self.printCurrentTask()
                self.currentProduct += 1

            self.wb.save(self.excelFilePath)
        print(f"Done!\nExcel File Path: {self.excelFilePath}")

    def getProductDatabaseLength(self):
        Session = sessionmaker(bind=self.productEngine)
        Base = declarative_base()

        Base.metadata.create_all(self.productEngine)
        session = Session()

        query = text("SELECT * FROM ProductInf WHERE isProcessed = 0")
        dataCount = len(session.execute(query).fetchall())
        self.productUnprocessedLength = dataCount
        self.currentProduct = 1


    def clearConsole(self):
        os.system('cls' if os.name == 'nt' else 'clear')    

    def printCurrentTask(self):
        self.clearConsole()
        print(f"Scraped datas are converting to product informations. (Insert to Excel File)")
        print("Started Time: ", self.startedTime)
        print("--"*20)
        print(f"{' ':<10}Product Counter")
        print(f"{' ':10}{self.currentProduct}/{self.productUnprocessedLength}")
        print("--"*20)
        print("Current Time: ", datetime.now().strftime("%H:%M:%S"))
